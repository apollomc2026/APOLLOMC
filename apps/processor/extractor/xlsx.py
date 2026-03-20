"""XLSX extraction module — extracts data from Excel spreadsheets."""

import io
from typing import Any

import openpyxl


def extract_xlsx(file_bytes: bytes) -> dict[str, Any]:
    """
    Extract structured content from an XLSX file.

    Returns:
        dict with keys: text, sheets, metadata
    """
    result: dict[str, Any] = {
        "text": "",
        "sheets": [],
        "metadata": {},
    }

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    result["metadata"] = {
        "sheet_names": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
    }

    full_text: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data: dict[str, Any] = {
            "name": sheet_name,
            "rows": [],
            "dimensions": ws.dimensions,
        }

        for row in ws.iter_rows(values_only=True):
            str_row = [str(cell) if cell is not None else "" for cell in row]
            sheet_data["rows"].append(str_row)
            non_empty = [c for c in str_row if c]
            if non_empty:
                full_text.append(" | ".join(non_empty))

        result["sheets"].append(sheet_data)

    result["text"] = "\n".join(full_text)

    return result
